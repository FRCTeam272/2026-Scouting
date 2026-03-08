import json
import os
import openpyxl
import requests
import normalize_functions

EXCEL_FILE = "HH.xlsx"
OUTPUT_DIR = "./data"
TBA_EVENT  = "2026pahat"
TBA_BASE   = "https://www.thebluealliance.com/api/v3"

def _tba_key():
    with open("secret") as f:
        return f.read().strip()

def tba_get(path):
    resp = requests.get(f"{TBA_BASE}{path}", headers={"X-TBA-Auth-Key": _tba_key()})
    resp.raise_for_status()
    return resp.json()

def enrich_with_tba(records, team_number):
    """Add TBA match keys, actual match numbers, and win/loss to each record."""
    team_key = f"frc{team_number}"
    try:
        matches = tba_get(f"/team/{team_key}/event/{TBA_EVENT}/matches/simple")
    except Exception as e:
        print(f"  TBA fetch failed for {team_key}: {e}")
        return records

    quals = sorted(
        [m for m in matches if m["comp_level"] == "qm"],
        key=lambda m: m["match_number"]
    )

    for i, record in enumerate(records):
        if i >= len(quals):
            break
        m = quals[i]
        red_teams  = m["alliances"]["red"]["team_keys"]
        blue_teams = m["alliances"]["blue"]["team_keys"]
        alliance   = "red" if team_key in red_teams else ("blue" if team_key in blue_teams else None)
        winning    = m.get("winning_alliance", "")
        if not winning:
            result = "tbd"
        elif alliance == winning:
            result = "win"
        else:
            result = "loss"

        record["_tba_match_num"] = m["match_number"]
        record["_tba_match_key"] = m["key"]
        record["_tba_result"]    = result
        record["_tba_url"]       = f"https://www.thebluealliance.com/match/{m['key']}"

    return records


def get_team_sheets(wb):
    """Return only sheets whose names start with a number."""
    return [name for name in wb.sheetnames if name[0].isdigit()]


def process_sheet(ws, sheet_name):
    """Process a single team sheet and return a list of match records."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = rows[1]  # Row 2 is the column headers
    data_rows = rows[2:]   # Row 3+ are match entries

    # Build index of columns that exist in coulmn_renames.
    # Duplicate Excel headers (same text twice) are handled by tracking already-used
    # column names: first occurrence uses the dict value; second occurrence of the same
    # header derives its name by replacing "Auto " with "Teleop " (handles the two
    # identical shooting accuracy headers in the sheet).
    col_map = []  # list of (col_index, renamed_column)
    used_col_names = set()
    for i, cell in enumerate(header_row):
        if cell not in normalize_functions.coulmn_renames:
            continue
        col_name = normalize_functions.coulmn_renames[cell]
        if col_name not in used_col_names:
            col_map.append((i, col_name))
            used_col_names.add(col_name)
        elif col_name.startswith("Auto "):
            # Second occurrence of a header whose first maps to "Auto X" → "Teleop X"
            teleop_name = "Teleop " + col_name[5:]
            col_map.append((i, teleop_name))
            used_col_names.add(teleop_name)
    

    records = []
    for row in data_rows:
        match_label = row[0]
        if not match_label:
            continue
        # Skip rows where all data columns are None or non-match rows like "Strengths"
        if all(v is None for v in row[1:]):
            continue
        if not str(match_label).strip().lower().startswith("match"):
            continue

        record = {"Match": str(match_label).strip()}
        for col_index, col_name in col_map:
            raw = row[col_index]
            if raw is None:
                record[col_name] = None
            else:
                try:
                  record[col_name] = normalize_functions.extract_key_mappings(col_name, str(raw))
                except Exception as e:
                  record[col_name] = None
                

        records.append(record)

    return records


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>HH Scouting 2026</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.2/dist/chart.umd.min.js"></script>
  <style>
    :root {
      --bg: #0f1117; --surface: #1a1d27; --surface2: #22263a; --border: #2e334d;
      --accent: #4f8ef7; --accent2: #f7a84f; --text: #e4e6f0; --muted: #7b82a0;
      --green: #4caf82; --red: #e05c5c; --yellow: #e0c45c; --purple: #9c6cf7; --teal: #4fc5cf;
      --sidebar-w: 260px;
    }
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body { font-family: 'Segoe UI', system-ui, sans-serif; background: var(--bg); color: var(--text); display: flex; flex-direction: column; height: 100vh; overflow: hidden; }

    /* Header */
    header { background: var(--surface); border-bottom: 1px solid var(--border); padding: 12px 16px; display: flex; align-items: center; gap: 12px; flex-shrink: 0; }
    header h1 { font-size: 1.1rem; font-weight: 700; color: var(--accent); letter-spacing: 0.04em; white-space: nowrap; }
    #team-count { color: var(--muted); font-size: 0.8rem; white-space: nowrap; }
    #search { margin-left: auto; background: var(--surface2); border: 1px solid var(--border); border-radius: 6px; padding: 6px 10px; color: var(--text); font-size: 0.85rem; width: 180px; outline: none; min-width: 0; }
    #search:focus { border-color: var(--accent); }
    #menu-btn { display: none; background: none; border: none; color: var(--text); cursor: pointer; padding: 4px; flex-shrink: 0; }

    /* Layout */
    .layout { display: flex; flex: 1; overflow: hidden; position: relative; }

    /* Sidebar */
    aside { width: var(--sidebar-w); flex-shrink: 0; background: var(--surface); border-right: 1px solid var(--border); display: flex; flex-direction: column; overflow: hidden; transition: transform 0.25s ease; }
    #sidebar-scroll { flex: 1; overflow-y: auto; padding-bottom: 8px; }
    .overview-btn { display: flex; align-items: center; gap: 8px; width: 100%; padding: 11px 16px; background: none; border: none; border-left: 3px solid transparent; border-bottom: 1px solid var(--border); color: var(--text); font-size: 0.9rem; font-weight: 600; cursor: pointer; }
    .overview-btn:hover { background: var(--surface2); }
    .overview-btn.active { background: var(--surface2); border-left-color: var(--accent2); color: var(--accent2); }

    /* Compare bar */
    #compare-bar { background: var(--surface2); border-bottom: 1px solid var(--border); padding: 8px 12px; display: none; flex-direction: column; gap: 6px; }
    #compare-chips { display: flex; flex-wrap: wrap; gap: 4px; }
    .chip { display: inline-flex; align-items: center; gap: 4px; background: var(--accent); color: #fff; border-radius: 12px; padding: 2px 8px; font-size: 0.75rem; font-weight: 600; }
    .chip button { background: none; border: none; color: #fff; cursor: pointer; font-size: 0.85rem; line-height: 1; padding: 0 0 0 2px; opacity: 0.8; }
    .chip button:hover { opacity: 1; }
    .compare-actions { display: flex; gap: 6px; }
    .compare-actions button { flex: 1; padding: 5px 8px; border-radius: 6px; border: none; cursor: pointer; font-size: 0.78rem; font-weight: 600; }
    #btn-do-compare { background: var(--accent); color: #fff; }
    #btn-do-compare:hover { opacity: 0.85; }
    #btn-clear-compare { background: var(--surface); color: var(--muted); border: 1px solid var(--border); }
    #btn-clear-compare:hover { color: var(--text); }

    /* Team items */
    .team-item { display: flex; align-items: center; gap: 8px; padding: 8px 12px 8px 8px; border-left: 3px solid transparent; }
    .team-item:hover { background: var(--surface2); }
    .team-item:hover .hide-btn { opacity: 1; }
    .team-item.active { background: var(--surface2); border-left-color: var(--accent); }
    .compare-check { width: 14px; height: 14px; flex-shrink: 0; accent-color: var(--accent); cursor: pointer; }
    .compare-check:disabled { opacity: 0.3; cursor: not-allowed; }
    .team-info { flex: 1; cursor: pointer; min-width: 0; }
    .team-number { font-weight: 700; font-size: 0.9rem; color: var(--accent); }
    .team-name { font-size: 0.75rem; color: var(--muted); margin-top: 1px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
    .team-badge { font-size: 0.68rem; background: var(--surface2); border: 1px solid var(--border); border-radius: 10px; padding: 1px 6px; color: var(--muted); white-space: nowrap; }
    .hide-btn { background: none; border: none; color: var(--muted); cursor: pointer; font-size: 1rem; line-height: 1; padding: 2px 4px; opacity: 0; border-radius: 4px; flex-shrink: 0; }
    .hide-btn:hover { color: var(--red); }
    #show-hidden-btn { display: none; width: 100%; padding: 10px 16px; background: none; border: none; border-top: 1px solid var(--border); color: var(--muted); font-size: 0.8rem; cursor: pointer; text-align: left; }
    #show-hidden-btn:hover { color: var(--text); background: var(--surface2); }

    /* Footer */
    footer { background: var(--surface); border-top: 1px solid var(--border); padding: 10px 20px; font-size: 0.75rem; color: var(--muted); text-align: center; flex-shrink: 0; }

    /* Overlay for mobile sidebar */
    #sidebar-overlay { display: none; position: fixed; inset: 0; background: rgba(0,0,0,0.5); z-index: 99; }

    /* Main */
    main { flex: 1; overflow-y: auto; padding: 20px; min-width: 0; }
    #placeholder { display: flex; flex-direction: column; align-items: center; justify-content: center; height: 100%; color: var(--muted); gap: 8px; }
    #placeholder svg { opacity: 0.2; }
    #team-view, #overview-view, #compare-view { display: none; }
    .section-title { font-size: 0.7rem; text-transform: uppercase; letter-spacing: 0.08em; color: var(--muted); margin-bottom: 12px; margin-top: 24px; }
    .section-title:first-child { margin-top: 0; }
    .page-header { margin-bottom: 20px; }
    .page-header h2 { font-size: 1.4rem; font-weight: 700; }
    .page-header h2 span { color: var(--accent); }
    .page-header p { color: var(--muted); font-size: 0.85rem; margin-top: 4px; }
    .tba-link { display: inline-flex; align-items: center; gap: 5px; color: var(--accent); font-size: 0.8rem; text-decoration: none; margin-top: 6px; }
    .tba-link:hover { text-decoration: underline; }
    .result-win  { color: var(--green); font-weight: 700; font-size: 0.75rem; margin-left: 6px; }
    .result-loss { color: var(--red);   font-weight: 700; font-size: 0.75rem; margin-left: 6px; }
    .result-tbd  { color: var(--muted); font-weight: 600; font-size: 0.75rem; margin-left: 6px; }
    .match-meta  { display: flex; align-items: center; gap: 8px; }
    .match-tba-link { font-size: 0.75rem; color: var(--muted); text-decoration: none; margin-left: auto; }
    .match-tba-link:hover { color: var(--accent); }

    /* Charts */
    .charts-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 16px; }
    .charts-grid.single { grid-template-columns: 1fr; }
    .chart-card { background: var(--surface); border: 1px solid var(--border); border-radius: 10px; padding: 14px 16px; }
    .chart-card h3 { font-size: 0.75rem; font-weight: 600; color: var(--muted); text-transform: uppercase; letter-spacing: 0.06em; margin-bottom: 12px; }
    .chart-wrap { position: relative; }
    .chart-wrap.tall { height: 300px; }
    .chart-wrap.medium { height: 200px; }
    .chart-wrap canvas { width: 100% !important; }
    .no-data { color: var(--muted); font-size: 0.85rem; font-style: italic; padding: 20px 0; }

    /* Match cards */
    .matches-grid { display: flex; flex-direction: column; gap: 14px; }
    .match-card { background: var(--surface); border: 1px solid var(--border); border-radius: 10px; overflow: hidden; }
    .match-card-header { background: var(--surface2); padding: 9px 14px; border-bottom: 1px solid var(--border); }
    .match-label { font-weight: 700; font-size: 0.9rem; color: var(--accent2); }
    .match-fields { display: grid; grid-template-columns: repeat(auto-fill, minmax(160px, 1fr)); gap: 1px; background: var(--border); }
    .field { background: var(--surface); padding: 8px 12px; }
    .field-label { font-size: 0.65rem; color: var(--muted); text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 2px; }
    .phase-badge { display:inline-block; font-size:0.55rem; font-weight:700; padding:1px 4px; border-radius:3px; margin-right:4px; vertical-align:middle; }
    .phase-auto   { background:#3b82f6; color:#fff; }
    .phase-teleop { background:#f97316; color:#fff; }
    .field-value { font-size: 0.88rem; font-weight: 600; }
    .val-null { color: var(--muted); font-style: italic; font-weight: 400; }
    .val-yes  { color: var(--green); }
    .val-no   { color: var(--red); }
    .val-na   { color: var(--muted); }
    .val-num  { color: var(--accent); }
    .val-str  { color: var(--text); }
    #error { display: none; color: var(--red); padding: 16px; background: var(--surface); border: 1px solid var(--red); border-radius: 8px; }

    /* Compare blocks */
    .compare-blocks { display: grid; grid-template-columns: repeat(auto-fill, minmax(300px, 1fr)); gap: 16px; }
    .compare-block { border: 1px solid var(--border); border-radius: 10px; overflow: hidden; }
    .compare-block-header { font-size: 0.9rem; font-weight: 700; color: var(--accent); padding: 9px 14px; background: var(--surface2); border-bottom: 1px solid var(--border); }

    /* ── Mobile ──────────────────────────────────────────────── */
    @media (max-width: 768px) {
      #menu-btn { display: block; }
      #search { width: 130px; }
      aside {
        position: fixed; top: 0; left: 0; height: 100%; z-index: 100;
        transform: translateX(-100%);
      }
      aside.open { transform: translateX(0); }
      #sidebar-overlay.visible { display: block; }
      main { padding: 14px; }
      .charts-grid { grid-template-columns: 1fr; }
      .chart-wrap.tall { height: 240px; }
      .chart-wrap.medium { height: 180px; }
      .match-fields { grid-template-columns: repeat(auto-fill, minmax(130px, 1fr)); }
      .compare-blocks { grid-template-columns: 1fr; }
    }
  </style>
</head>
<body>
<header>
  <button id="menu-btn" onclick="toggleSidebar()" aria-label="Menu">
    <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
      <line x1="3" y1="6" x2="21" y2="6"/><line x1="3" y1="12" x2="21" y2="12"/><line x1="3" y1="18" x2="21" y2="18"/>
    </svg>
  </button>
  <h1>HH Scouting 2026</h1>
  <span id="team-count"></span>
  <input id="search" type="text" placeholder="Search..." />
</header>
<div id="sidebar-overlay" onclick="toggleSidebar()"></div>
<div class="layout">
  <aside id="sidebar">
    <div id="compare-bar">
      <div id="compare-chips"></div>
      <div class="compare-actions">
        <button id="btn-do-compare" onclick="showComparison()">Compare (<span id="compare-count">0</span>)</button>
        <button id="btn-clear-compare" onclick="clearCompare()">Clear</button>
      </div>
    </div>
    <div id="sidebar-scroll">
      <button class="overview-btn" id="overview-btn" onclick="showOverview()">
        <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2">
          <rect x="3" y="3" width="7" height="7"/><rect x="14" y="3" width="7" height="7"/>
          <rect x="3" y="14" width="7" height="7"/><rect x="14" y="14" width="7" height="7"/>
        </svg>
        Overview
      </button>
      <div id="team-list"></div>
    </div>
    <button id="show-hidden-btn" onclick="showAllHidden()"></button>
  </aside>
  <main>
    <div id="placeholder">
      <svg width="56" height="56" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.5">
        <path d="M9 5H7a2 2 0 00-2 2v12a2 2 0 002 2h10a2 2 0 002-2V7a2 2 0 00-2-2h-2"/>
        <rect x="9" y="3" width="6" height="4" rx="1"/>
        <path d="M9 12h6M9 16h4"/>
      </svg>
      <p>Select a team or view Overview</p>
    </div>
    <div id="error"></div>
    <div id="team-view"></div>
    <div id="overview-view"></div>
    <div id="compare-view"></div>
  </main>
</div>
<footer>Data Collected and Organized by FRC 708 &nbsp;·&nbsp; Visualizations provided by FRC 272</footer>
<script>
const SCOUTING_EMBEDDED = __SCOUTING_DATA__;

const dataCache = {};
let allTeams = [];
const activeCharts = {};
let hiddenSet = new Set(JSON.parse(localStorage.getItem('hhscout_hidden') || '[]'));
let compareSet = new Set();

Chart.defaults.color = '#7b82a0';
Chart.defaults.borderColor = '#2e334d';
Chart.defaults.font.family = "'Segoe UI', system-ui, sans-serif";

const PALETTE = {
  blue: '#4f8ef7', orange: '#f7a84f', green: '#4caf82',
  red: '#e05c5c', purple: '#9c6cf7', teal: '#4fc5cf', yellow: '#e0c45c',
};
const ACTION_COLORS  = [PALETTE.green, PALETTE.purple, PALETTE.teal, PALETTE.red];
const INACTIVE_COLORS = [PALETTE.red, PALETTE.teal, PALETTE.yellow];

function parsePercent(v) {
  if (v === null || v === undefined) return null;
  if (typeof v === 'number') return v;
  const n = parseInt(v, 10); return isNaN(n) ? null : n;
}
function valueClass(val) {
  if (val === null || val === undefined) return 'val-null';
  if (val === 'Y') return 'val-yes'; if (val === 'N') return 'val-no';
  if (val === 'N/A') return 'val-na'; if (typeof val === 'number') return 'val-num';
  return 'val-str';
}
function displayValue(val) { return (val === null || val === undefined) ? '—' : val; }
function avg(arr) { const v = arr.filter(x => x !== null); return v.length ? v.reduce((a,b)=>a+b,0)/v.length : null; }
function makeChart(id, config) {
  if (activeCharts[id]) { activeCharts[id].destroy(); delete activeCharts[id]; }
  const el = document.getElementById(id); if (!el) return;
  activeCharts[id] = new Chart(el, config);
}
function baseOptions() {
  return { responsive: true, maintainAspectRatio: false,
    plugins: { legend: { labels: { boxWidth: 12, padding: 12, font: { size: 11 } } },
               tooltip: { backgroundColor: '#1a1d27', borderColor: '#2e334d', borderWidth: 1, titleColor: '#e4e6f0', bodyColor: '#7b82a0', padding: 10 } },
    scales: { x: { grid: { color: '#2e334d' }, ticks: { color: '#7b82a0' } },
               y: { grid: { color: '#2e334d' }, ticks: { color: '#7b82a0' } } } };
}
function doughnutOpts(pos='right') {
  return { responsive: true, maintainAspectRatio: false,
    plugins: { legend: { position: pos, labels: { boxWidth: 12, padding: 10, font: { size: 11 } } },
               tooltip: { backgroundColor: '#1a1d27', borderColor: '#2e334d', borderWidth: 1, titleColor: '#e4e6f0', bodyColor: '#7b82a0', padding: 10 } } };
}
function makeDoughnut(id, labels, values, colors) {
  makeChart(id, { type: 'doughnut', data: { labels, datasets: [{
    data: values,
    backgroundColor: colors.map(c => c+'cc'), borderColor: colors, borderWidth: 1,
  }]}, options: doughnutOpts() });
}
function getRecords(f) { return dataCache[f] || []; }

function hideAll() {
  ['placeholder','team-view','overview-view','compare-view','error'].forEach(id => document.getElementById(id).style.display = 'none');
  document.querySelectorAll('.team-item').forEach(el => el.classList.remove('active'));
  document.getElementById('overview-btn').classList.remove('active');
}

function teamStats(t) {
  const records = getRecords(t.file);
  const [num] = t.team.split(' - ');
  const autoAccVals   = records.map(r => parsePercent(r['Auto Shooting Accuracy %'])).filter(v => v !== null);
  const teleopAccVals = records.map(r => parsePercent(r['Teleop Shooting Accuracy %'])).filter(v => v !== null);
  const climbY = l => records.filter(r => r[l] === 'Y').length;
  const climbN = l => records.filter(r => r[l] === 'N').length;
  const climbRate = l => { const tot = climbY(l)+climbN(l); return tot > 0 ? climbY(l)/tot*100 : null; };
  const autoActions   = { Scoring: 0, Feeding: 0, Collecting: 0, 'Did Nothing': 0 };
  const activeRoles   = { Scoring: 0, Feeding: 0, Collecting: 0, Defense: 0 };
  const inactiveRoles = { Defense: 0, Feeder: 0, 'Did Nothing': 0 };
  records.forEach(r => {
    if (r['Auto Actions'] && r['Auto Actions'] in autoActions) autoActions[r['Auto Actions']]++;
    [r['Active Shift 1 Actions'],   r['Active Shift 2 Actions']  ].forEach(v => { if (v && v in activeRoles)   activeRoles[v]++; });
    [r['Inactive Shift 1 Actions'], r['Inactive Shift 2 Actions']].forEach(v => { if (v && v in inactiveRoles) inactiveRoles[v]++; });
  });
  return { num, file: t.file, team: t.team, records, avgAuto: avg(autoAccVals), avgTeleop: avg(teleopAccVals), climbRate, autoActions, activeRoles, inactiveRoles };
}

// ── Sidebar toggle (mobile) ───────────────────────────────────────────────
function toggleSidebar() {
  const sb = document.getElementById('sidebar');
  const ov = document.getElementById('sidebar-overlay');
  const open = sb.classList.toggle('open');
  ov.classList.toggle('visible', open);
}
function closeSidebar() {
  document.getElementById('sidebar').classList.remove('open');
  document.getElementById('sidebar-overlay').classList.remove('visible');
}

// ── Compare ───────────────────────────────────────────────────────────────
function toggleCompare(filename, checked) {
  if (checked && compareSet.size < 6) compareSet.add(filename); else compareSet.delete(filename);
  updateCompareBar(); updateCheckboxStates();
}
function removeFromCompare(filename) {
  compareSet.delete(filename);
  const cb = document.querySelector(`.compare-check[data-file="${filename}"]`);
  if (cb) cb.checked = false;
  updateCompareBar(); updateCheckboxStates();
}
function clearCompare() {
  compareSet.clear();
  document.querySelectorAll('.compare-check').forEach(cb => { cb.checked = false; cb.disabled = false; });
  updateCompareBar();
}
function updateCompareBar() {
  const bar = document.getElementById('compare-bar');
  if (compareSet.size === 0) { bar.style.display = 'none'; return; }
  bar.style.display = 'flex';
  document.getElementById('compare-chips').innerHTML = allTeams.filter(t => compareSet.has(t.file)).map(t =>
    `<span class="chip">${t.team.split(' - ')[0]} <button onclick="removeFromCompare('${t.file}')">×</button></span>`).join('');
  document.getElementById('compare-count').textContent = compareSet.size;
}
function updateCheckboxStates() {
  document.querySelectorAll('.compare-check').forEach(cb => { cb.disabled = !cb.checked && compareSet.size >= 6; });
}

// ── Hide ──────────────────────────────────────────────────────────────────
function hideTeam(e, filename) {
  e.stopPropagation();
  hiddenSet.add(filename); compareSet.delete(filename);
  localStorage.setItem('hhscout_hidden', JSON.stringify([...hiddenSet]));
  updateCompareBar(); rebuildSidebar();
}
function showAllHidden() {
  hiddenSet.clear(); localStorage.setItem('hhscout_hidden', JSON.stringify([])); rebuildSidebar();
}
function rebuildSidebar() {
  const q = document.getElementById('search').value.toLowerCase();
  buildSidebar(allTeams.filter(t => !hiddenSet.has(t.file) && t.team.toLowerCase().includes(q)));
  const btn = document.getElementById('show-hidden-btn');
  btn.style.display = hiddenSet.size > 0 ? 'block' : 'none';
  if (hiddenSet.size > 0) btn.textContent = `Show ${hiddenSet.size} hidden team${hiddenSet.size !== 1 ? 's' : ''}`;
}

// ── Team View ─────────────────────────────────────────────────────────────
function loadTeam(entry) {
  hideAll(); closeSidebar();
  saveNav({view: 'team', file: entry.file});
  document.querySelector(`.team-item[data-file="${entry.file}"]`)?.classList.add('active');

  const records = getRecords(entry.file);
  const [number, ...nameParts] = entry.team.split(' - ');
  const name = nameParts.join(' - ');

  const tbaUrl        = entry.tba_url || '';
  const statboticsUrl = entry.statbotics_url || '';
  const extLink = (href, label) => `<a class="tba-link" href="${href}" target="_blank">
    <svg width="12" height="12" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M18 13v6a2 2 0 01-2 2H5a2 2 0 01-2-2V8a2 2 0 012-2h6"/><polyline points="15 3 21 3 21 9"/><line x1="10" y1="14" x2="21" y2="3"/></svg>
    ${label}</a>`;
  const tv = document.getElementById('team-view');
  tv.innerHTML = `
    <div class="page-header">
      <h2><span>#${number}</span> ${name}</h2>
      <p>${records.length} match${records.length !== 1 ? 'es' : ''} scouted</p>
      <div style="display:flex;gap:16px;flex-wrap:wrap;margin-top:4px;">
        ${tbaUrl        ? extLink(tbaUrl,        'The Blue Alliance') : ''}
        ${statboticsUrl ? extLink(statboticsUrl, 'Statbotics')        : ''}
      </div>
    </div>
    <div class="charts-grid">
      <div class="chart-card"><h3>Active Shift Distribution</h3><div class="chart-wrap medium"><canvas id="c-active"></canvas></div></div>
      <div class="chart-card"><h3>Inactive Shift Distribution</h3><div class="chart-wrap medium"><canvas id="c-inactive"></canvas></div></div>
    </div>
    <div class="charts-grid single">
      <div class="chart-card"><h3>Shooting Accuracy %</h3><div class="chart-wrap medium"><canvas id="c-acc"></canvas></div></div>
    </div>
    <p class="section-title">Match Records</p>
    <div class="matches-grid">
      ${records.map(r => {
        const matchLabel = r._tba_match_num ? `Qual ${r._tba_match_num}` : r.Match;
        const resultCls  = r._tba_result === 'win' ? 'result-win' : r._tba_result === 'loss' ? 'result-loss' : 'result-tbd';
        const resultText = r._tba_result === 'win' ? 'WIN' : r._tba_result === 'loss' ? 'LOSS' : r._tba_result === 'tbd' ? 'TBD' : '';
        const tbaMatchLink = r._tba_url ? `<a class="match-tba-link" href="${r._tba_url}" target="_blank">TBA &#8599;</a>` : '';
        const fields = Object.entries(r).filter(([k]) => !k.startsWith('_') && k !== 'Match').map(([k,v]) => {
          let labelHtml;
          if (k === 'Auto Shooting Accuracy %')   labelHtml = `<span class="phase-badge phase-auto">AUTO</span>Shooting Accuracy %`;
          else if (k === 'Teleop Shooting Accuracy %') labelHtml = `<span class="phase-badge phase-teleop">TELEOP</span>Shooting Accuracy %`;
          else labelHtml = k;
          return `<div class="field"><div class="field-label">${labelHtml}</div><div class="field-value ${valueClass(v)}">${displayValue(v)}</div></div>`;
        }).join('');
        return `<div class="match-card">
          <div class="match-card-header">
            <div class="match-meta">
              <span class="match-label">${matchLabel}</span>
              ${resultText ? `<span class="${resultCls}">${resultText}</span>` : ''}
              ${tbaMatchLink}
            </div>
          </div>
          <div class="match-fields">${fields}</div>
        </div>`;
      }).join('')}
    </div>`;
  tv.style.display = 'block';

  // Shooting accuracy
  const matchLabels = records.map(r => r.Match.replace('Match ', 'M'));
  const autoAcc   = records.map(r => parsePercent(r['Auto Shooting Accuracy %']));
  const teleopAcc = records.map(r => parsePercent(r['Teleop Shooting Accuracy %']));
  if (autoAcc.some(v => v !== null) || teleopAcc.some(v => v !== null)) {
    makeChart('c-acc', { type: 'bar', data: { labels: matchLabels, datasets: [
      { label: 'Auto %',   data: autoAcc,   backgroundColor: PALETTE.blue  +'cc', borderColor: PALETTE.blue,   borderWidth: 1 },
      { label: 'Teleop %', data: teleopAcc, backgroundColor: PALETTE.orange+'cc', borderColor: PALETTE.orange, borderWidth: 1 },
    ]}, options: { ...baseOptions(), scales: { x: { grid:{color:'#2e334d'}, ticks:{color:'#7b82a0'} }, y: { grid:{color:'#2e334d'}, ticks:{color:'#7b82a0'}, min:0, max:100 } }}});
  } else {
    document.getElementById('c-acc').parentElement.innerHTML = '<p class="no-data">No shooting accuracy data recorded</p>';
  }

  // Active shift distribution
  const s = teamStats(entry);
  const activeTotal = Object.values(s.activeRoles).reduce((a,b)=>a+b,0);
  if (activeTotal > 0) {
    makeDoughnut('c-active', Object.keys(s.activeRoles), Object.values(s.activeRoles), ACTION_COLORS);
  } else {
    document.getElementById('c-active').parentElement.innerHTML = '<p class="no-data">No active shift data recorded</p>';
  }

  // Inactive shift distribution
  const inactiveTotal = Object.values(s.inactiveRoles).reduce((a,b)=>a+b,0);
  if (inactiveTotal > 0) {
    makeDoughnut('c-inactive', Object.keys(s.inactiveRoles), Object.values(s.inactiveRoles), INACTIVE_COLORS);
  } else {
    document.getElementById('c-inactive').parentElement.innerHTML = '<p class="no-data">No inactive shift data recorded</p>';
  }
}

// ── Comparison View ───────────────────────────────────────────────────────
function showComparison() {
  if (compareSet.size < 2) return;
  hideAll(); closeSidebar();
  saveNav({view: 'compare', files: [...compareSet]});
  const data   = allTeams.filter(t => compareSet.has(t.file)).map(teamStats);
  const labels = data.map(d => d.num);

  const cv = document.getElementById('compare-view');
  cv.innerHTML = `
    <div class="page-header">
      <h2>Comparing <span>${data.length} Teams</span></h2>
      <p>${labels.join(' · ')}</p>
    </div>
    <div class="charts-grid">
      <div class="chart-card"><h3>Avg Shooting Accuracy %</h3><div class="chart-wrap medium"><canvas id="cmp-acc"></canvas></div></div>
      <div class="chart-card"><h3>Climb Success Rate %</h3><div class="chart-wrap medium"><canvas id="cmp-climb"></canvas></div></div>
    </div>
    <div class="charts-grid">
      <div class="chart-card"><h3>Auto Action Distribution %</h3><div class="chart-wrap medium"><canvas id="cmp-auto"></canvas></div></div>
      <div class="chart-card"><h3>Active Shift Distribution %</h3><div class="chart-wrap medium"><canvas id="cmp-active"></canvas></div></div>
    </div>
    <div class="charts-grid single">
      <div class="chart-card"><h3>Inactive Shift Distribution %</h3><div class="chart-wrap medium"><canvas id="cmp-inactive"></canvas></div></div>
    </div>
    <p class="section-title">Match Records</p>
    <div class="compare-blocks">
      ${data.map(d => `
        <div class="compare-block">
          <div class="compare-block-header">#${d.num} · ${d.records.length} match${d.records.length!==1?'es':''}</div>
          ${d.records.map(r => {
            const matchLabel = r._tba_match_num ? `Qual ${r._tba_match_num}` : r.Match;
            const resultCls  = r._tba_result==='win'?'result-win':r._tba_result==='loss'?'result-loss':'result-tbd';
            const resultText = r._tba_result==='win'?'WIN':r._tba_result==='loss'?'LOSS':r._tba_result==='tbd'?'TBD':'';
            const tbaLink    = r._tba_url ? `<a class="match-tba-link" href="${r._tba_url}" target="_blank">TBA &#8599;</a>` : '';
            const fields = Object.entries(r).filter(([k])=>!k.startsWith('_')&&k!=='Match').map(([k,v])=>
              `<div class="field"><div class="field-label">${k}</div><div class="field-value ${valueClass(v)}">${displayValue(v)}</div></div>`).join('');
            return `<div class="match-card" style="border-radius:0;border-left:none;border-right:none;border-top:none">
              <div class="match-card-header"><div class="match-meta"><span class="match-label">${matchLabel}</span>${resultText?`<span class="${resultCls}">${resultText}</span>`:''} ${tbaLink}</div></div>
              <div class="match-fields">${fields}</div></div>`;
          }).join('')}
        </div>`).join('')}
    </div>`;
  cv.style.display = 'block';

  const stackedOpts = { ...baseOptions(), scales: {
    x: { grid:{color:'#2e334d'}, ticks:{color:'#7b82a0'}, stacked:true },
    y: { grid:{color:'#2e334d'}, ticks:{color:'#7b82a0'}, stacked:true, min:0, max:100 } }};

  makeChart('cmp-acc', { type:'bar', data:{ labels, datasets:[
    { label:'Avg Auto %',   data:data.map(d=>d.avgAuto  !==null?Math.round(d.avgAuto)  :null), backgroundColor:PALETTE.blue  +'cc', borderColor:PALETTE.blue,   borderWidth:1 },
    { label:'Avg Teleop %', data:data.map(d=>d.avgTeleop!==null?Math.round(d.avgTeleop):null), backgroundColor:PALETTE.orange+'cc', borderColor:PALETTE.orange, borderWidth:1 },
  ]}, options:{...baseOptions(), scales:{ x:{grid:{color:'#2e334d'},ticks:{color:'#7b82a0'}}, y:{grid:{color:'#2e334d'},ticks:{color:'#7b82a0'},min:0,max:100} }}});

  makeChart('cmp-climb', { type:'bar', data:{ labels, datasets:[
    { label:'L1 %', data:data.map(d=>d.climbRate('Climb L1')??0), backgroundColor:PALETTE.green +'cc', borderColor:PALETTE.green,  borderWidth:1 },
    { label:'L2 %', data:data.map(d=>d.climbRate('Climb L2')??0), backgroundColor:PALETTE.yellow+'cc', borderColor:PALETTE.yellow, borderWidth:1 },
    { label:'L3 %', data:data.map(d=>d.climbRate('Climb L3')??0), backgroundColor:PALETTE.purple+'cc', borderColor:PALETTE.purple, borderWidth:1 },
  ]}, options:{...baseOptions(), scales:{ x:{grid:{color:'#2e334d'},ticks:{color:'#7b82a0'}}, y:{grid:{color:'#2e334d'},ticks:{color:'#7b82a0'},min:0,max:100} }}});

  const stackedBar = (id, keys, colors, getter) => makeChart(id, { type:'bar', data:{ labels, datasets: keys.map((k,i) => ({
    label:k, backgroundColor:colors[i]+'cc', borderColor:colors[i], borderWidth:1,
    data: data.map(d => { const tot=Object.values(getter(d)).reduce((a,b)=>a+b,0); return tot?Math.round(getter(d)[k]/tot*100):0; }),
  }))}, options: stackedOpts });

  stackedBar('cmp-auto',     ['Scoring','Feeding','Collecting','Did Nothing'], ACTION_COLORS,   d => d.autoActions);
  stackedBar('cmp-active',   ['Scoring','Feeding','Collecting','Defense'],     ACTION_COLORS,   d => d.activeRoles);
  stackedBar('cmp-inactive', ['Defense','Feeder','Did Nothing'],               INACTIVE_COLORS, d => d.inactiveRoles);
}

// ── Overview ──────────────────────────────────────────────────────────────
function showOverview() {
  hideAll(); closeSidebar();
  saveNav({view: 'overview'});
  document.getElementById('overview-btn').classList.add('active');

  const teamData  = allTeams.map(teamStats);
  const accSorted = [...teamData].filter(d => d.avgAuto!==null||d.avgTeleop!==null)
    .sort((a,b) => ((b.avgAuto||0)+(b.avgTeleop||0)) - ((a.avgAuto||0)+(a.avgTeleop||0)));
  const gAuto     = { Scoring:0, Feeding:0, Collecting:0, 'Did Nothing':0 };
  const gActive   = { Scoring:0, Feeding:0, Collecting:0, Defense:0 };
  const gInactive = { Defense:0, Feeder:0, 'Did Nothing':0 };
  teamData.forEach(d => {
    Object.entries(d.autoActions).forEach(([k,v])   => { gAuto[k]     += v; });
    Object.entries(d.activeRoles).forEach(([k,v])   => { gActive[k]   += v; });
    Object.entries(d.inactiveRoles).forEach(([k,v]) => { gInactive[k] += v; });
  });

  const ov = document.getElementById('overview-view');
  ov.innerHTML = `
    <div class="page-header"><h2>Overview</h2><p>Aggregated data across all ${allTeams.length} teams</p></div>
    <p class="section-title">Shooting Accuracy</p>
    <div class="charts-grid">
      <div class="chart-card"><h3>Avg Auto Accuracy % — by Team</h3><div class="chart-wrap tall"><canvas id="ov-auto-acc"></canvas></div></div>
      <div class="chart-card"><h3>Avg Teleop Accuracy % — by Team</h3><div class="chart-wrap tall"><canvas id="ov-teleop-acc"></canvas></div></div>
    </div>
    <p class="section-title">Climb Performance</p>
    <div class="charts-grid single">
      <div class="chart-card"><h3>Climb Success Rate % (L1 / L2 / L3) — by Team</h3><div class="chart-wrap tall"><canvas id="ov-climb"></canvas></div></div>
    </div>
    <p class="section-title">Action Distribution — All Teams</p>
    <div class="charts-grid">
      <div class="chart-card"><h3>Auto Action Mix</h3><div class="chart-wrap medium"><canvas id="ov-auto-dist"></canvas></div></div>
      <div class="chart-card"><h3>Active Shift Mix</h3><div class="chart-wrap medium"><canvas id="ov-active-dist"></canvas></div></div>
    </div>
    <div class="charts-grid single">
      <div class="chart-card"><h3>Inactive Shift Mix</h3><div class="chart-wrap medium"><canvas id="ov-inactive-dist"></canvas></div></div>
    </div>`;
  ov.style.display = 'block';

  const hbarOpts = (color) => ({ ...baseOptions(), indexAxis:'y', scales: {
    x: { grid:{color:'#2e334d'}, ticks:{color:'#7b82a0'}, min:0, max:100 },
    y: { grid:{color:'transparent'}, ticks:{color:'#7b82a0', font:{size:10}} } }});

  const autoAccTeams = accSorted.filter(d => d.avgAuto !== null);
  makeChart('ov-auto-acc', { type:'bar', data:{ labels:autoAccTeams.map(d=>d.num), datasets:[{
    label:'Avg Auto %', data:autoAccTeams.map(d=>Math.round(d.avgAuto)),
    backgroundColor:PALETTE.blue+'cc', borderColor:PALETTE.blue, borderWidth:1,
  }]}, options: hbarOpts() });

  const teleopAccTeams = accSorted.filter(d => d.avgTeleop !== null);
  if (teleopAccTeams.length > 0) {
    makeChart('ov-teleop-acc', { type:'bar', data:{ labels:teleopAccTeams.map(d=>d.num), datasets:[{
      label:'Avg Teleop %', data:teleopAccTeams.map(d=>Math.round(d.avgTeleop)),
      backgroundColor:PALETTE.orange+'cc', borderColor:PALETTE.orange, borderWidth:1,
    }]}, options: hbarOpts() });
  } else {
    document.getElementById('ov-teleop-acc').parentElement.innerHTML = '<p class="no-data">No teleop accuracy data available</p>';
  }

  const climbTeams = teamData.filter(d => d.climbRate('Climb L1')!==null||d.climbRate('Climb L2')!==null||d.climbRate('Climb L3')!==null);
  if (climbTeams.length > 0) {
    makeChart('ov-climb', { type:'bar', data:{ labels:climbTeams.map(d=>d.num), datasets:[
      { label:'L1 %', data:climbTeams.map(d=>d.climbRate('Climb L1')), backgroundColor:PALETTE.green +'cc', borderColor:PALETTE.green,  borderWidth:1 },
      { label:'L2 %', data:climbTeams.map(d=>d.climbRate('Climb L2')), backgroundColor:PALETTE.yellow+'cc', borderColor:PALETTE.yellow, borderWidth:1 },
      { label:'L3 %', data:climbTeams.map(d=>d.climbRate('Climb L3')), backgroundColor:PALETTE.purple+'cc', borderColor:PALETTE.purple, borderWidth:1 },
    ]}, options:{...baseOptions(), scales:{ x:{grid:{color:'#2e334d'},ticks:{color:'#7b82a0',maxRotation:45,font:{size:10}}}, y:{grid:{color:'#2e334d'},ticks:{color:'#7b82a0'},min:0,max:100} }}});
  } else {
    document.getElementById('ov-climb').parentElement.innerHTML = '<p class="no-data">No climb data available</p>';
  }

  makeDoughnut('ov-auto-dist',     Object.keys(gAuto),     Object.values(gAuto),     ACTION_COLORS);
  makeDoughnut('ov-active-dist',   Object.keys(gActive),   Object.values(gActive),   ACTION_COLORS);
  makeDoughnut('ov-inactive-dist', Object.keys(gInactive), Object.values(gInactive), INACTIVE_COLORS);
}

// ── Sidebar ───────────────────────────────────────────────────────────────
function buildSidebar(teams) {
  document.getElementById('team-list').innerHTML = teams.map(t => {
    const [number, ...nameParts] = t.team.split(' - ');
    const name = nameParts.join(' - ');
    const safe = JSON.stringify(t).replace(/"/g, '&quot;');
    const checked  = compareSet.has(t.file) ? 'checked' : '';
    const disabled = !compareSet.has(t.file) && compareSet.size >= 6 ? 'disabled' : '';
    return `<div class="team-item" data-file="${t.file}">
      <input type="checkbox" class="compare-check" data-file="${t.file}" ${checked} ${disabled}
             onclick="event.stopPropagation(); toggleCompare('${t.file}', this.checked)">
      <div class="team-info" onclick="loadTeam(${safe})">
        <div class="team-number">${number}</div>
        <div class="team-name">${name}</div>
      </div>
      <span class="team-badge">${t.matches}m</span>
      <button class="hide-btn" title="Hide team" onclick="hideTeam(event,'${t.file}')">×</button>
    </div>`;
  }).join('');
}

document.getElementById('search').addEventListener('input', () => rebuildSidebar());

// ── State persistence ─────────────────────────────────────────────────────
function saveNav(state) {
  localStorage.setItem('hhscout_nav', JSON.stringify(state));
}

function restoreNav() {
  let state;
  try { state = JSON.parse(localStorage.getItem('hhscout_nav')); } catch { return; }
  if (!state) return;

  if (state.view === 'overview') {
    showOverview();
  } else if (state.view === 'team' && state.file) {
    const entry = allTeams.find(t => t.file === state.file);
    if (entry) loadTeam(entry);
  } else if (state.view === 'compare' && state.files?.length >= 2) {
    state.files.forEach(f => { if (allTeams.find(t => t.file === f)) compareSet.add(f); });
    updateCompareBar();
    updateCheckboxStates();
    rebuildSidebar();
    showComparison();
  }
}

// Boot
allTeams = SCOUTING_EMBEDDED.index;
Object.assign(dataCache, SCOUTING_EMBEDDED.teams);
document.getElementById('team-count').textContent = `${allTeams.length} teams`;
rebuildSidebar();
restoreNav();
</script>
</body>
</html>
"""


def build_html(index, all_records):
    embedded = json.dumps({"index": index, "teams": all_records}, separators=(",", ":"))
    return HTML_TEMPLATE.replace("__SCOUTING_DATA__", embedded)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    team_sheets = get_team_sheets(wb)
    print(f"Found {len(team_sheets)} team sheets")

    index = []
    all_records = {}
    for name in team_sheets:
        team_number = name.split(" - ")[0].strip()
        records = process_sheet(wb[name], name)
        records = enrich_with_tba(records, team_number)
        filename = name.replace(" ", "_").replace("/", "-") + ".json"
        filepath = os.path.join(OUTPUT_DIR, filename)
        with open(filepath, "w") as f:
            json.dump(records, f, indent=2)
        print(f"Wrote {len(records)} records -> {filepath}")
        tba_team_url = f"https://www.thebluealliance.com/team/{team_number}"
        statbotics_url = f"https://www.statbotics.io/team/{team_number}"
        index.append({"team": name, "file": filename, "matches": len(records), "tba_url": tba_team_url, "statbotics_url": statbotics_url})
        all_records[filename] = records

    with open(os.path.join(OUTPUT_DIR, "index.json"), "w") as f:
        json.dump(index, f, indent=2)
    print(f"Wrote index.json with {len(index)} teams")

    html_path = "index.html"
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(build_html(index, all_records))
    print(f"Wrote {html_path}")

    wb.close()


if __name__ == "__main__":
    main()
